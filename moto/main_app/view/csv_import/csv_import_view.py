import string, random
from datetime import datetime
from django.contrib.auth import login as auth_login, authenticate, logout as auth_logout
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User, Group
from django.contrib import messages

from django.shortcuts import redirect
from django.shortcuts import render
from django.http import HttpResponse

from openpyxl import Workbook, load_workbook

from main_app.models import Nutzer, Personal, Raum, Gruppe, AG, Schueler, Zeitraum, AGZeit, Datumsraum, AGKategorie 
import re

@login_required(redirect_field_name="login")
def csv_import_view(request):
    if request.user.is_superuser:
        if request.method == 'POST':
            excel_file=None
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
            response['Content-Disposition'] = 'attachment; filename={date}-otps.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

            try:
                excel_file = request.FILES['excel_file']
            except:
                messages.error(request, "Keine Datei Hochgeladen")
                return redirect('csv_import')
            
            if(excel_file.name.endswith(".xlsx")):              # check if xlsx datei

                # try:
                        wb = load_workbook(excel_file, read_only=True)  # save workbook as Excel Workbook (*.xlsx) and not as Strict Open XML Spreadsheet (*.xlsx)
                                # Parameter setzten
                        
                        fehler_tabelle=''
                        optionlist_reset = request.POST.getlist('checkbox_database')

                        optionlist = request.POST.getlist('checkbox_sheet')

                        if 'option_room' in optionlist:
                            if 'option_overwrite' in optionlist_reset:
                                Raum.objects.all().delete()

                            wss = wb['Raeume']
                            for row in wss.iter_rows(min_row=2, values_only=True):
                                try:
                                    raum_nr, geschoss, kapazitaet, kategorie, hex_farbe = row

                                    # Überprüfen, ob Raum bereits existiert
                                    if not Raum.objects.filter(raum_nr=raum_nr).exists():
                                        # Standardwerte setzen, falls nötig
                                        kategorie = kategorie if kategorie else "Allgemein"
                                        
                                        # Hex-Farbe überprüfen
                                        if hex_farbe:
                                            if hex_farbe and not is_valid_hex_color(hex_farbe):
                                                messages.error(request, f"Ungültiger Hex-Code für Raum {raum_nr}: {hex_farbe}")
                                                error = True
                                                continue 

                                        color = hex_farbe if hex_farbe else "#FFFFFF"

                                        Raum.objects.create(
                                            raum_nr=raum_nr,
                                            geschoss=geschoss,
                                            kapazitaet=kapazitaet,
                                            kategorie=kategorie,
                                            color=color
                                        )
                                except Exception as e:
                                    error = True
                                    messages.error(request, f"Fehler bei der Verarbeitung von Raum {row}: {e}")

                        if 'option_user' in optionlist:
                            if 'option_overwrite' in optionlist_reset:
                                for p in Personal.objects.all():
                                    nutzer = p.nutzer
                                    p.delete()
                                    nutzer.delete()
                                for user in User.objects.all():
                                    if not user.username == "root":
                                        user.delete()
                            fehler_tabelle='Fehler in Tabelle Personal: '
                            wb_otp = Workbook()
                            activ_sheet = wb_otp.active
                            activ_sheet.title = "OTP"
                            activ_sheet.append(['Username', 'Einmal Passwort'])
                            wsp = wb['Personal']
                            for index, row in enumerate(wsp.iter_rows(min_row=2, values_only=True)):              # Erstellung Personal
                                
                                error = False
                                vorname, nachname, funktion, rechte = row
                                new_nutzer = Nutzer.objects.create(vorname=vorname,nachname=nachname)
                                try:
                                    if rechte=='':
                                        rechte = 'Ohne Rolle'
                                    rechte_gruppe = Group.objects.get(name=rechte)
                                except:
                                    error=True
                                    messages.error(request, fehler_tabelle+"rechte_gruppe "+str(rechte)+" vom Personal "+str(vorname)+" "+str(nachname)+" in Zeile " + str(index) +" existiert nicht. Options: (Admin, Gruppenleitung, Raumbetreuer, Ohne Rolle)")
                                randompw = ''.join(random.choice(string.ascii_letters+string.digits) for _ in range(6))     #erstellung random passwort
                                username = (str(vorname)+str(nachname)).lower()
                                zahl = 0
                                is_username_unique = False
                                while not is_username_unique:
                                    un2 = username
                                    if(zahl > 0):
                                        un2 = username + str(zahl)
                                    if (User.objects.filter(username=un2).exists()==True):
                                        zahl += 1
                                    else:
                                        username=un2
                                        is_username_unique=True
                                # print(username)
                                # print(randompw)
                                activ_sheet.append([username, randompw])
                                newuser = User.objects.create_user(username=username, password=randompw)
                                newuser.groups.add(rechte_gruppe)
                                if(rechte_gruppe.name=='Admin'):
                                    newuser.is_superuser = True
                                newuser.save()
                                if not error:
                                    Personal.objects.create(rolle=funktion, nutzer=new_nutzer, user=newuser, rechte_gruppe=rechte_gruppe)
        
                            wb_otp.save(response)

                        if 'option_ag' in optionlist:
                            fehler_tabelle='Fehler in Tabelle AGs: '
                            if 'option_overwrite' in optionlist_reset:
                                AG.objects.all().delete()
                            wss = wb['AGs']
                            for index, row in enumerate(wss.iter_rows(min_row=2, values_only=True)):              # Erstellung Nutzer
                                error = False
                                name, ag_kategorie, max_anzahl, offene_AG, ag_leiter, montag, dienstag, mittwoch, donnerstag, freitag, angebotsstart, angebotsende = row

                                zahl = 1
                                is_name_unique = False
                                while not is_name_unique:
                                    un2 = name
                                    if(zahl > 1):
                                        un2 = name + str(zahl)
                                    if (AG.objects.filter(name=un2).exists()==True):
                                        zahl += 1
                                    else:
                                        name=un2
                                        is_name_unique=True
                                try:
                                    user = User.objects.get(username=ag_leiter)
                                    leiter = Personal.objects.get(user=user)
                                except:
                                    error=True
                                    messages.error(request, fehler_tabelle + "ag_leiter " + str(name)+" in Zeile " + str(index) +" existiert nicht.")
                                is_offene_AG = False
                                try:
                                    ag_kategorie = AGKategorie.objects.get(name=ag_kategorie)
                                except:
                                    error=True
                                    messages.error(request, fehler_tabelle+"ag_kategorie "+str(max_anzahl)+" in Zeile " + str(index) +" existiert nicht.")
                                try:
                                    max_anzahl = int(max_anzahl)
                                except:
                                    error=True
                                    messages.error(request, fehler_tabelle+"max_anzahl "+str(max_anzahl)+" in Zeile " + str(index) +" muss eine Zahl sein.")
                                if(type(angebotsstart)==datetime and type(angebotsende)==datetime):                           
                                    datumsraum = Datumsraum.objects.create(startdatum=angebotsstart, enddatum=angebotsende)
                                else:
                                    error = True
                                    messages.error(request, fehler_tabelle+"Datum in Zeile " + str(index) +" kann nicht Formatiert werden. Format: dd.mm.YYYY")                         
                                if(offene_AG.lower()=='true' or offene_AG.lower()=='ja'):
                                        is_offene_AG = True
                                if not error:        
                                    ag = AG.objects.create(name=name,ag_kategorie=ag_kategorie,max_anzahl=max_anzahl,offene_AG=is_offene_AG, leiter=leiter, angebots_datum_raum=datumsraum)
                                    
                                    # TODO: Error Handling
                                    if not montag==None:
                                        montag = montag.replace(" ","")
                                        montag = montag.split(";")
                                        create_ag_zeiten(montag, AGZeit.WOCHENTAG.MONTAG, ag)

                                    if not dienstag == None:
                                        dienstag = dienstag.replace(" ","")
                                        dienstag = dienstag.split(";")
                                        create_ag_zeiten(dienstag, AGZeit.WOCHENTAG.DIENSTAG, ag)

                                    if not mittwoch == None:
                                        mittwoch = mittwoch.replace(" ","")
                                        mittwoch = mittwoch.split(";")
                                        create_ag_zeiten(mittwoch, AGZeit.WOCHENTAG.MITTWOCH, ag)

                                    if not donnerstag == None:
                                        donnerstag = donnerstag.replace(" ","")
                                        donnerstag = donnerstag.split(";")
                                        create_ag_zeiten(donnerstag, AGZeit.WOCHENTAG.DONNERSTAG, ag)

                                    if not freitag == None:
                                        freitag = freitag.replace(" ","")
                                        freitag = freitag.split(";")
                                        create_ag_zeiten(freitag, AGZeit.WOCHENTAG.FREITAG, ag)

                        if 'option_group' in optionlist:
                            fehler_tabelle='Fehler in Tabelle Gruppen: '
                            if 'option_overwrite' in optionlist_reset:
                                Gruppe.objects.all().delete()
                            wss = wb['Gruppen']
                            for index, row in enumerate(wss.iter_rows(min_row=2, values_only=True)):              # Erstellung Nutzer
                                #print(row)c
                                error = False
                                name, gruppen_leiter, raum = row
                                gl_l = gruppen_leiter.replace(" ", "")
                                gl_l = gl_l.split(',')
                                if(Gruppe.objects.filter(name=name).exists()==False):
                                    try:   
                                        aufsichtspersonen = []
                                        for gl_name in gl_l:
                                            try:
                                                user = (User.objects.get(username=gl_name))
                                                gruppen_leiter = Personal.objects.get(user=user)
                                                aufsichtspersonen.append(gruppen_leiter)
                                            except:
                                                error=True
                                                messages.error(request, fehler_tabelle+"gruppen_leiter "+str(gl_name)+" in Zeile " + str(index) +" existiert nicht.")
                                    except:
                                        error=True
                                        messages.error(request, fehler_tabelle+"Fehler bei erstellen der Gruppe "+str(name)+" in Zeile " + str(index) +" existiert nicht.")
                                    try:
                                        raum = Raum.objects.get(raum_nr=raum)
                                    except:
                                        error=True
                                        messages.error(request, fehler_tabelle+"raum "+ str(raum) +" in Zeile " + str(index) +" existiert nicht.")
                                    if not error:
                                        neue_gruppe = Gruppe.objects.create(name=name, raum=raum)
                                        neue_gruppe.gruppen_leiter.set(aufsichtspersonen)
                                        neue_gruppe.save()
                                else:
                                    messages.error(request, fehler_tabelle+"Gruppe "+ str(name)+" in Zeile " + str(index) +" existiert bereits.")

                        if 'option_pupil' in optionlist:
                            fehler_tabelle='Fehler in Tabelle Schueler: '
                            if 'option_overwrite' in optionlist_reset:
                                for s in Schueler.objects.all():
                                    nutzer = s.user_id
                                    s.delete()
                                    print(nutzer.vorname + " " + nutzer.nachname)
                                    nutzer.delete()
                            wss = wb['Schueler']
                            for index, row in enumerate(wss.iter_rows(min_row=2, values_only=True)):              # Erstellung Nutzer
                                
                                error = False
                                vorname, nachname, gruppen_name, klasse, bus_kind, name_eb, kontakt_eb  = row
                                if(bus_kind.lower()=='true' or bus_kind.lower()=='ja'):
                                        bus_kind = True
                                else:
                                    bus_kind = False
                                try:
                                    gruppen_id = Gruppe.objects.get(name=gruppen_name)
                                except:
                                    error=True
                                    messages.error(request, fehler_tabelle+"Gruppe "+str(gruppen_id)+" in Zeile " + str(index) +" existiert nicht.")
                                new_nutzer = Nutzer.objects.create(vorname=vorname,nachname=nachname)
                                schueler = Schueler.objects.create(klasse=klasse, bus_kind=bus_kind, name_eb=name_eb, kontakt_eb=kontakt_eb, user_id=new_nutzer,gruppen_id=gruppen_id)
                                schueler.save()

                        if 'option_user' in optionlist:
                            return response                            # Site after Download
                        return redirect('csv_import')
                    # return response           # Weiterleitung wenn alles funktioniert hat
                # except:
                #     messages.error(request, "Fehler bei importieren der Excel datei. Bitte überprüfe ob die Datei als 'Excel Arbeitsmappe (*.xlsx)' gespeichter ist, alle Tabellen die importier oder zurückgesetzt werden sollen existieren und ob die Spalten, wie in der Vorlage, richtig bennant sind.")              
                #     return redirect('master_web')      # Weiterleitung bei falscher xlsx datei
                                    
            print('Falscher Dateityp')        
            return render(request, 'csv_import/csv_import.html')      # Weiterleitung bei flaschen datei Typen
        
        return render(request, 'csv_import/csv_import.html')
    else:
        return redirect("csv_import")
    

def create_ag_zeiten(list, day, ag):

    for uhrzeit in list:
        if not uhrzeit == "":
            uhrzeit = uhrzeit.split("-")
            if(len(uhrzeit[0].split(":"))>1):
                startzeit_stunde = uhrzeit[0].split(":")[0]
                startzeit_minute = uhrzeit[0].split(":")[1]
            else:
                startzeit_stunde = uhrzeit[0]
                startzeit_minute = "00"
            if(len(uhrzeit[1].split(":"))>1):
                endzeit_stunde = uhrzeit[1].split(":")[0]
                endzeit_minute = uhrzeit[1].split(":")[1]
            else:
                endzeit_stunde = uhrzeit[1]
                endzeit_minute = "00"

            startzeit = datetime.strptime(startzeit_stunde+':'+startzeit_minute, '%H:%M').time()
            endzeit = datetime.strptime(endzeit_stunde+':'+endzeit_minute, '%H:%M').time()
            zeitraum = Zeitraum.objects.create(startzeit=startzeit, endzeit=endzeit)
            agzeit = AGZeit.objects.create(wochentag=day, zeitraum=zeitraum)
            ag.ag_zeit.add(agzeit)
    
def is_valid_hex_color(color):
    return bool(re.fullmatch(r"^#[0-9A-Fa-f]{6}$", color))
